import random
import colorama
from openpyxl import Workbook,load_workbook
from time import sleep
from colorama import Fore, Back, Style

colorama.init()

wb = load_workbook("words.xlsx")
ws = wb.active

print(Fore.BLUE, '\nQUIZ UYGULAMASINA HOŞGELDİNİZ..\n', Style.RESET_ALL)

giris = input("İLK KÜTÜPHANE => '1',\nİKİNCİ KÜTÜPHANE => '2',\nSEÇİM: ")

if giris == '1':
    ws = wb['first2514']

    soruSayisi = int(input('SORU SAYISINI SEÇİN: '))

    for i in range(soruSayisi):

        sleep(2)

        rand_row = random.randint(1, 1128)

        istedigimiz_satir = ws[rand_row]

        kelimenin_ingilizcesi = istedigimiz_satir[0].value
        kelimenin_okunusu = istedigimiz_satir[1].value
        kelimenin_turkcesi = istedigimiz_satir[2].value

        x = input("\nTÜRKÇE => İNGİLİZCE 'T',\nİNGİLİZCE => TÜRKÇE 'E',\nÇIKIŞ => 'Q' GİRİN.\nSEÇİM: ")

        if x == 'E':
            print(Fore.LIGHTBLUE_EX, rand_row, '---', kelimenin_ingilizcesi.upper(), Style.RESET_ALL)

            a = input("\nTÜRKÇESİ => 'T',\nOKUNUŞU => 'P' GİRİN.\nSEÇİM: ")

            if a == 'T':
                print(Fore.LIGHTRED_EX, rand_row, '---', kelimenin_turkcesi.upper(), Style.RESET_ALL)
            elif a == 'P':
                print(Fore.LIGHTYELLOW_EX, rand_row, '---', kelimenin_okunusu.upper(), Style.RESET_ALL)
            else:
                print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)
        elif x == 'T':
            print(Fore.LIGHTRED_EX, rand_row, '---', kelimenin_turkcesi.upper(), Style.RESET_ALL)

            b = input("\nİNGİLİZCESİ => 'E',\nOKUNUŞU => 'P' GİRİN.\nSEÇİM: ")

            if b == 'E':
                print(Fore.LIGHTBLUE_EX, rand_row, '---', kelimenin_ingilizcesi.upper(), Style.RESET_ALL)
            elif b == 'P':
                print(Fore.LIGHTYELLOW_EX, rand_row, '---', kelimenin_okunusu.upper(), Style.RESET_ALL)
            else:
                print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)
        elif x == 'Q':
            print(Fore.LIGHTGREEN_EX, '\nÇIKIŞ YAPILIYOR.', Style.RESET_ALL)
            break
        else:
            print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)

if giris == '2':
    ws = wb['second1127']

    soruSayisi = int(input('SORU SAYISINI SEÇİN: '))

    for i in range(soruSayisi):

        sleep(2)

        rand_row = random.randint(1, 1128)

        istedigimiz_satir = ws[rand_row]

        kelimenin_ingilizcesi = istedigimiz_satir[0].value
        kelimenin_okunusu = istedigimiz_satir[1].value
        kelimenin_turkcesi = istedigimiz_satir[2].value

        x = input("\nTÜRKÇE => İNGİLİZCE 'T',\nİNGİLİZCE => TÜRKÇE 'E',\nÇIKIŞ => 'Q' GİRİN.\nSEÇİM: ")

        if x == 'E':
            print(Fore.LIGHTBLUE_EX, rand_row, '---', kelimenin_ingilizcesi.upper(), Style.RESET_ALL)

            a = input("\nTÜRKÇESİ => 'T',\nOKUNUŞU => 'P' GİRİN.\nSEÇİM: ")

            if a == 'T':
                print(Fore.LIGHTRED_EX, rand_row, '---', kelimenin_turkcesi.upper(), Style.RESET_ALL)
            elif a == 'P':
                print(Fore.LIGHTYELLOW_EX, rand_row, '---', kelimenin_okunusu.upper(), Style.RESET_ALL)
            else:
                print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)
        elif x == 'T':
            print(Fore.LIGHTRED_EX, rand_row, '---', kelimenin_turkcesi.upper(), Style.RESET_ALL)

            b = input("\nİNGİLİZCESİ => 'E',\nOKUNUŞU => 'P' GİRİN.\nSEÇİM: ")

            if b == 'E':
                print(Fore.LIGHTBLUE_EX, rand_row, '---', kelimenin_ingilizcesi.upper(), Style.RESET_ALL)
            elif b == 'P':
                print(Fore.LIGHTYELLOW_EX, rand_row, '---', kelimenin_okunusu.upper(), Style.RESET_ALL)
            else:
                print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)
        elif x == 'Q':
            print(Fore.LIGHTGREEN_EX, '\nÇIKIŞ YAPILIYOR.', Style.RESET_ALL)
            break
        else:
            print(Fore.CYAN, "\nGEÇERSİZ BİR SEÇİM YAPTINIZ.", Style.RESET_ALL)
